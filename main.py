from openpyxl import load_workbook 
from openpyxl.cell.read_only import EmptyCell
import pprint
pp = pprint.PrettyPrinter(indent=2, width=1)


class XlTex(object):
  """docstring for XlTex"""
  def __init__(self, wb_path):
    super(XlTex, self).__init__()
    self.wb = load_workbook(filename=wb_path, read_only=True)

  def gera_marcador_invertido(self, marcador_str):
    # gera o marcador de abertura/fechamento correspondente ao informado
    assert type(marcador_str) == str

    marcador_str_strip = marcador_str.strip()
    if marcador_str_strip.startswith('{{'):
      return f"{marcador_str_strip.split(' ')[0].replace('{', '')}}}}}"
    if marcador_str_strip.endswith('}}'):
      return f"{{{{{marcador_str_strip.split(' ')[-1].replace('}', '')}"
    assert False


  def pega_marcacoes(self):
    """
    retorna dicionario com listas
    de celulas marcadas para planilha
    """
    
    # dicionario que vai armazenar uma lista contendo as celulas marcadas
    marcacoes_por_planilha = dict()

    for sheet_name in self.wb.sheetnames:
      ws = self.wb[sheet_name]
      marcacoes_por_planilha[ws] = list()

      # localiza as celulas com marcações
      for row in ws.rows:
        for cell in row:
          if not isinstance(cell, EmptyCell):
            if type(cell.value) == str and (cell.value.startswith(r'{{') or cell.value.endswith(r'}}')):
              marcacoes_por_planilha[ws].append(cell)
    return marcacoes_por_planilha

  def pega_limites(self, marcacoes_por_planilha):
    """
    retorna lista com tuplas de um ou dois elementos
    contendo as celulas limite das regiões a serem exportadas
    """
    assert type(marcacoes_por_planilha) == dict
    
    limites = list()
    
    for sheet_name in self.wb.sheetnames:
      ws = self.wb[sheet_name]
      for cell in marcacoes_por_planilha[ws]:
        # as marcações estão em uma mesma celula
        if cell.value.startswith(r'{{') and cell.value.endswith(r'}}'):
          limites.append((cell,))
          continue
        if cell.value.startswith(r'{{'):
          marcador_invertido = self.gera_marcador_invertido(cell.value)
          celula_fechamento_list = [x for x in marcacoes_por_planilha[ws] if x.value.endswith(marcador_invertido)]
          try:
            celula_fechamento = celula_fechamento_list[0]
          except Exception as e:
            print(f"Não foi encontr a célula de fechamento para `{cell.value}`")
            raise e
          
          limites.append((cell, celula_fechamento,))

    return limites

  def pega_regioes(self, limites):
    """
    returna lista contendo matrizes das regiões
    """
    assert type(limites) == list

    regioes = list()
    for limite in limites:
      if limite[0]== limite[-1]:
        #regioes.append([[limite[0]]])
        continue
      
      ws = limite[0].parent
      # inicializa matriz
      mat = list()
      # filtra as regiões. Será que dá para fazer melhor?
      for row_i, row in enumerate(ws.rows):
        if row_i >= (limite[0].row-1) and row_i < limite[-1].row:
          mat_row = list()
          for column_i, cell in enumerate(row):
            if column_i >= limite[0].column and column_i < (limite[-1].column-1):
              mat_row.append(cell)
          mat.append(mat_row)
      regioes.append(mat)
    return regioes


def main():
  xlTex = XlTex('teste.xlsx')
  marcacoes_por_planilha = xlTex.pega_marcacoes()
  limites = xlTex.pega_limites(marcacoes_por_planilha)
  regioes = xlTex.pega_regioes(limites)
  pp.pprint(regioes)
  #wb.close()

if __name__ == '__main__':
  main()